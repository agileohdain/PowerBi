#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extracteur de données  donnees.xlsx  ->  bloc DATA JavaScript (grain MENSUEL).

Pourquoi ce script existe
-------------------------
Le skill `powerbi-prototype` embarque les données DANS le HTML (grain mensuel +
séries par dimension + masques d'activité + agrégats catégoriels) pour afficher
l'année N et rendre la variation vs N-1 possible. Dériver ce modèle « à la
main » mène à des KPI mal interprétés et des dimensions oubliées. Ce script
**auto-détecte** le modèle de n'importe quel Excel (star/snowflake : une table de
faits datée + dimensions) et émet un contrat **normalisé**, indépendant du domaine
(cyclisme, RH, ventes, finance…).

Deux modes
----------
* **Défaut (normalisé, multi-domaine)** : auto-détection -> `const DATA = {…}`
  avec FACTS / BY_DIM / DIM_COUNTS / CATEGORY_COUNTS / ACTIVE_MASKS / SCALARS /
  META. Aucun nom de feuille/colonne n'est codé en dur.
* **`--profile cyclisme`** : émet le contrat *legacy* (consts `KM`, `RIDES`,
  `USURE_STATUT`, `USER_MASKS`, `KM_PAYS_M`, `KM_MARQUE_M`, …) utilisé par les
  maquettes cyclisme existantes (Veloh, agiledss). C'est un **contrat de
  données** (format des `const` émises), **pas** un layout de maquette à
  copier. Compat descendante.

Manifeste (override, optionnel)
-------------------------------
L'auto-détection propose un manifeste JSON sur **stderr**. Copiez-le dans
`clients/<client>/data-manifest.json` pour reprendre la main (forcer la table de
faits, la colonne date, les mesures/dimensions retenues, l'entité active). Si le
fichier existe, il est utilisé **verbatim** (plus aucune devinette). Aucune
dépendance externe : JSON uniquement.

Usage :
  python extract-data.py clients/<client>/donnees.xlsx
  python extract-data.py clients/<client>/donnees.xlsx --profile cyclisme
  python extract-data.py clients/<client>/donnees.xlsx --manifest clients/<client>/data-manifest.json

Dépendance : openpyxl  (pip install openpyxl)
"""
import sys
import os
import re
import json
import argparse
import datetime
from collections import defaultdict, Counter, deque

try:
    import openpyxl
except ImportError:
    sys.stderr.write("ERREUR: openpyxl manquant -> pip install openpyxl\n")
    sys.exit(2)

CAT_MAX = 40           # cardinalité max pour qu'une colonne texte soit une dimension
PERSON_RE = re.compile(r"(?i)utilisateur|user|employe|employee|client|person|agent|membre|collab")
ID_RE = re.compile(r"(?i)^id(_|$)|_id$")
NAME_RE = re.compile(r"(?i)nom|prenom|name|email|mail|tel|phone|adresse|address")


# ---------------------------------------------------------------------------
# Chargement & profilage
# ---------------------------------------------------------------------------
def is_date(v):
    return isinstance(v, (datetime.datetime, datetime.date))


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def is_id(name):
    return bool(ID_RE.search(name))


def load_wb(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        try:
            hdr = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        rows = []
        for r in it:
            if not any(v is not None for v in r):
                continue
            rows.append({hdr[i]: r[i] for i in range(len(hdr))})
        sheets[name] = {"header": hdr, "rows": rows}
    return sheets


def classify(name, values):
    vals = [v for v in values if v is not None]
    if not vals:
        return "empty", 0
    if is_id(name):
        return "id", len(set(vals))
    n = len(vals)
    ndate = sum(1 for v in vals if is_date(v))
    nnum = sum(1 for v in vals if is_num(v))
    nstr = sum(1 for v in vals if isinstance(v, str))
    distinct = len(set(vals))
    if ndate >= n * 0.8:
        return "date", distinct
    if nnum >= n * 0.8:
        return "numeric", distinct
    if nstr >= n * 0.8 and not NAME_RE.search(name):
        if distinct <= CAT_MAX and distinct < n * 0.9:
            return "categorical", distinct
        return "text", distinct
    return "mixed", distinct


def profile(sheets):
    """Type de chaque colonne de chaque feuille + clé primaire détectée."""
    info = {}
    for name, sh in sheets.items():
        hdr, rows = sh["header"], sh["rows"]
        cols = {}
        for c in hdr:
            if not c:
                continue
            t, distinct = classify(c, [r.get(c) for r in rows])
            cols[c] = {"type": t, "distinct": distinct}
        # clé primaire = 1re colonne id et unique
        pk = None
        if hdr:
            first = hdr[0]
            if first in cols and cols[first]["type"] == "id" and \
               cols[first]["distinct"] == len(rows) and len(rows) > 0:
                pk = first
        info[name] = {"cols": cols, "pk": pk, "n": len(rows)}
    return info


# ---------------------------------------------------------------------------
# Auto-détection du modèle
# ---------------------------------------------------------------------------
def detect_fact(info):
    cands = []
    for name, s in info.items():
        has_date = any(c["type"] == "date" for c in s["cols"].values())
        n_meas = sum(1 for c in s["cols"].values() if c["type"] == "numeric")
        if has_date and n_meas >= 1:
            cands.append((s["n"], n_meas, name))
    if not cands:
        return None
    cands.sort(reverse=True)
    return cands[0][2]


def pick_date_col(info, fact):
    dates = [c for c, d in info[fact]["cols"].items() if d["type"] == "date"]
    for c in dates:  # préférence à un nom "date" pur
        if c.strip().lower() in ("date", "jour", "day", "date_"):
            return c
    return dates[0] if dates else None


def build_join_graph(info):
    """col -> feuille dont elle est la PK (uniquement si non ambiguë)."""
    pk_sheets = defaultdict(list)
    for name, s in info.items():
        if s["pk"]:
            pk_sheets[s["pk"]].append(name)
    return {col: names[0] for col, names in pk_sheets.items() if len(names) == 1}


def resolve_dims(info, fact):
    """BFS depuis la faits à travers le graphe de jointures (gère les ponts).
    Retourne la liste des dimensions : {name, via_sheet, path, value_col}."""
    col_to_pk = build_join_graph(info)
    visited = {fact}
    queue = deque([(fact, [])])   # (feuille courante, chemin [(col, sheet), ...])
    dims = []
    while queue:
        sheet, path = queue.popleft()
        # dimensions catégorielles de la feuille courante
        for col, d in info[sheet]["cols"].items():
            if d["type"] == "categorical":
                dims.append({"name": col, "via_sheet": sheet, "path": path, "value_col": col})
        # poursuite du graphe via les clés étrangères
        for col, d in info[sheet]["cols"].items():
            if d["type"] != "id" or col == info[sheet]["pk"]:
                continue
            tgt = col_to_pk.get(col)
            if tgt and tgt not in visited:
                visited.add(tgt)
                queue.append((tgt, path + [(col, tgt)]))
    return dims, visited


def resolve_value(fact_row, path, value_col, rows_by_pk):
    cur = fact_row
    for col, sheet in path:
        key = cur.get(col)
        if key is None:
            return None
        cur = rows_by_pk.get(sheet, {}).get(key)
        if cur is None:
            return None
    return cur.get(value_col)


# ---------------------------------------------------------------------------
# Construction normalisée
# ---------------------------------------------------------------------------
def build_normalized(sheets, info, manifest=None):
    manifest = manifest or {}
    fact = manifest.get("fact_sheet") or detect_fact(info)
    if not fact or fact not in sheets:
        raise SystemExit("ERREUR: aucune table de faits datée trouvée "
                         "(feuille avec une colonne date + une mesure numérique).")
    date_col = manifest.get("date_col") or pick_date_col(info, fact)
    fact_rows = sheets[fact]["rows"]

    # mesures = colonnes numériques non-id de la faits
    measures = [c for c, d in info[fact]["cols"].items() if d["type"] == "numeric"]
    if manifest.get("measures"):
        measures = [m for m in measures if m in manifest["measures"]]

    # grain mensuel continu
    dates = [r.get(date_col) for r in fact_rows if is_date(r.get(date_col))]
    if not dates:
        raise SystemExit("ERREUR: colonne date '%s' sans date valide dans '%s'." % (date_col, fact))
    months, midx = month_index(dates)
    N = len(months)

    # lookups pk -> ligne pour chaque feuille
    rows_by_pk = {}
    for name, sh in sheets.items():
        pk = info[name]["pk"]
        if pk:
            rows_by_pk[name] = {r.get(pk): r for r in sh["rows"]}

    dims, visited = resolve_dims(info, fact)
    if manifest.get("dims"):
        dims = [d for d in dims if d["name"] in manifest["dims"]]

    # agrégation mensuelle
    FACTS = {m: [0.0] * N for m in measures}
    COUNT = [0] * N
    BY_DIM = {}
    for d in dims:
        BY_DIM[d["name"]] = {}
    for r in fact_rows:
        dt = r.get(date_col)
        if not is_date(dt):
            continue
        i = midx[(dt.year, dt.month)]
        COUNT[i] += 1
        for m in measures:
            v = r.get(m)
            if is_num(v):
                FACTS[m][i] += float(v)
        for d in dims:
            val = resolve_value(r, d["path"], d["value_col"], rows_by_pk)
            if val is None:
                continue
            slot = BY_DIM[d["name"]].setdefault(str(val), {m: [0.0] * N for m in measures} | {"_count": [0] * N})
            slot["_count"][i] += 1
            for m in measures:
                v = r.get(m)
                if is_num(v):
                    slot[m][i] += float(v)
    FACTS["_count"] = COUNT
    for m in measures:
        FACTS[m] = [round(x, 1) for x in FACTS[m]]
    for d in BY_DIM.values():
        for slot in d.values():
            for m in measures:
                slot[m] = [round(x, 1) for x in slot[m]]

    # DIM_COUNTS : comptage des lignes de la table de dimension (statique)
    DIM_COUNTS = {}
    for d in dims:
        sh = sheets[d["via_sheet"]]["rows"]
        cnt = Counter(str(r.get(d["value_col"])) for r in sh if r.get(d["value_col"]) is not None)
        DIM_COUNTS[d["name"]] = dict(sorted(cnt.items(), key=lambda kv: -kv[1]))

    # CATEGORY_COUNTS : feuilles NON atteintes depuis la faits (ex. usure)
    CATEGORY_COUNTS = {}
    for name, sh in sheets.items():
        if name == fact or name in visited:
            continue
        for col, dd in info[name]["cols"].items():
            if dd["type"] == "categorical":
                cnt = Counter(str(r.get(col)) for r in sh["rows"] if r.get(col) is not None)
                if cnt:
                    CATEGORY_COUNTS.setdefault(name, {})[col] = dict(sorted(cnt.items(), key=lambda kv: -kv[1]))

    # ACTIVE_MASKS : entité « personne » atteinte depuis la faits. On ne retient
    # que les feuilles porteuses de dimensions catégorielles (via_sheet) : un pont
    # 100 % ID (ex. ASSOC_*) ne peut pas être une entité métier. Itération sur la
    # liste `dims` (ordre BFS déterministe), jamais sur le set `visited`.
    ACTIVE_MASKS = {}
    activity_sheet = manifest.get("activity_entity")
    if not activity_sheet:
        seen = set()
        cand = [s for s in (d["via_sheet"] for d in dims) if not (s in seen or seen.add(s))]
        for s in cand:
            if s != fact and PERSON_RE.search(s):
                activity_sheet = s
                break
    if activity_sheet and activity_sheet in visited:
        pk = info[activity_sheet]["pk"]
        # chemin vers cette feuille (retrouvé via les dims)
        apath = None
        for d in dims:
            if d["via_sheet"] == activity_sheet:
                apath = d["path"]
                break
        if apath is None:
            apath = find_path(info, fact, activity_sheet)
        if pk and apath is not None:
            for r in fact_rows:
                dt = r.get(date_col)
                if not is_date(dt):
                    continue
                ent = resolve_value(r, apath, pk, rows_by_pk)
                if ent is None:
                    continue
                ACTIVE_MASKS[ent] = ACTIVE_MASKS.get(ent, 0) | (1 << midx[(dt.year, dt.month)])
            ACTIVE_MASKS = {str(k): v for k, v in ACTIVE_MASKS.items()}

    # SCALARS
    SCALARS = {}
    for name, sh in sheets.items():
        SCALARS["NB_" + re.sub(r"\W+", "_", name).upper()] = len(sh["rows"])
    for dname, cnt in DIM_COUNTS.items():
        SCALARS["NB_" + re.sub(r"\W+", "_", dname).upper()] = len(cnt)
    for name, sh in sheets.items():
        for col, dd in info[name]["cols"].items():
            if dd["type"] == "numeric" and name != fact:
                vals = [r.get(col) for r in sh["rows"] if is_num(r.get(col))]
                if vals:
                    SCALARS["AVG_" + re.sub(r"\W+", "_", col).upper()] = round(sum(vals) / len(vals), 1)
            # DISTINCT_ : comptage des valeurs distinctes d'une colonne id
            # (ex. table pont ASSOC_* -> vélos/usagers distincts attribués).
            if dd["type"] == "id":
                vals = [r.get(col) for r in sh["rows"] if r.get(col) is not None]
                if vals:
                    SCALARS["DISTINCT_" + re.sub(r"\W+", "_", name + "_" + col).upper()] = len(set(vals))

    data = {
        "N": N,
        "MONTH_META": [{"year": y, "month": m, "quarter": (m - 1) // 3 + 1} for y, m in months],
        "FACTS": FACTS,
        "BY_DIM": BY_DIM,
        "DIM_COUNTS": DIM_COUNTS,
        "CATEGORY_COUNTS": CATEGORY_COUNTS,
        "ACTIVE_MASKS": ACTIVE_MASKS,
        "SCALARS": SCALARS,
        "META": {
            "fact_sheet": fact,
            "date_col": date_col,
            "measures": measures,
            "dims": [{"name": d["name"], "via_sheet": d["via_sheet"]} for d in dims],
            "activity_entity": activity_sheet,
            "category_sources": {k: list(v.keys()) for k, v in CATEGORY_COUNTS.items()},
        },
    }
    proposed = {
        "fact_sheet": fact, "date_col": date_col, "measures": measures,
        "dims": [d["name"] for d in dims], "activity_entity": activity_sheet,
    }
    return data, proposed


def find_path(info, fact, target):
    col_to_pk = build_join_graph(info)
    visited = {fact}
    queue = deque([(fact, [])])
    while queue:
        sheet, path = queue.popleft()
        if sheet == target:
            return path
        for col, d in info[sheet]["cols"].items():
            if d["type"] != "id" or col == info[sheet]["pk"]:
                continue
            tgt = col_to_pk.get(col)
            if tgt and tgt not in visited:
                visited.add(tgt)
                queue.append((tgt, path + [(col, tgt)]))
    return None


def month_index(dates):
    ym = sorted({(d.year, d.month) for d in dates})
    (y0, m0), (y1, m1) = ym[0], ym[-1]
    months = []
    y, m = y0, m0
    while (y, m) <= (y1, m1):
        months.append((y, m))
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
    return months, {t: i for i, t in enumerate(months)}


# ---------------------------------------------------------------------------
# Profil legacy « cyclisme » (compat descendante avec les maquettes existantes)
# ---------------------------------------------------------------------------
def build_cyclisme(sheets):
    def rows(name):
        return sheets.get(name, {"rows": []})["rows"]
    users = rows("DIM_UTILISATEUR")
    velos = rows("DIM_VELO")
    assoc = rows("ASSOC_UTILISATEUR_VELO")
    sorties = rows("FAIT_SORTIES")
    usure = rows("FAIT_USURE_COMPOSANT")
    if not sorties:
        raise SystemExit("ERREUR: FAIT_SORTIES vide ou absent (profil cyclisme).")

    dates = [r["DATE"] for r in sorties if r.get("DATE")]
    months, idx = month_index(dates)
    N = len(months)
    KM = [0.0] * N
    RIDES = [0] * N
    MINUTES = [0.0] * N
    active = [set() for _ in range(N)]
    uvel_velo = {a["ID_UTILISATEUR_VELO"]: a["ID_Velo"] for a in assoc}
    uvel_uid = {a["ID_UTILISATEUR_VELO"]: a["ID_Utilisateur"] for a in assoc}
    velo_marque = {v["ID_Velo"]: v["Marque"] for v in velos}
    uid_pays = {u["ID_Utilisateur"]: u["Pays"] for u in users}
    uid_ville = {u["ID_Utilisateur"]: u["Ville"] for u in users}
    KM_PAYS = defaultdict(lambda: [0.0] * N)
    RIDES_PAYS = defaultdict(lambda: [0] * N)
    KM_MARQUE = defaultdict(lambda: [0.0] * N)
    KM_VILLE = defaultdict(lambda: [0.0] * N)
    for r in sorties:
        d = r["DATE"]
        i = idx[(d.year, d.month)]
        km = float(r.get("NB_KM") or 0)
        mn = float(r.get("MINUTES") or 0)
        KM[i] += km
        RIDES[i] += 1
        MINUTES[i] += mn
        uvel = r.get("ID_UTILISATEUR_VELO")
        uid = uvel_uid.get(uvel)
        velo = uvel_velo.get(uvel)
        if uid is not None:
            active[i].add(uid)
            KM_PAYS[uid_pays[uid]][i] += km
            RIDES_PAYS[uid_pays[uid]][i] += 1
            KM_VILLE[uid_ville[uid]][i] += km
        if velo is not None:
            KM_MARQUE[velo_marque[velo]][i] += km
    rnd = lambda a: [round(x, 1) for x in a]
    KM, MINUTES = rnd(KM), rnd(MINUTES)
    for d in (KM_PAYS, KM_MARQUE, KM_VILLE):
        for k in d:
            d[k] = rnd(d[k])

    USER_MASKS = {}
    for uid in sorted(uid_pays):
        mask = 0
        for i in range(N):
            if uid in active[i]:
                mask |= (1 << i)
        if mask:
            USER_MASKS[str(uid)] = mask

    sort_by_total = lambda d: dict(sorted(d.items(), key=lambda kv: -sum(kv[1])))
    assigned = {a["ID_Velo"] for a in assoc}
    out = {
        "KM_LABELS": [f"{y}-{m:02d}" for y, m in months],
        "MONTH_META": [{"year": y, "month": m, "quarter": (m - 1) // 3 + 1} for y, m in months],
        "KM": KM, "RIDES": RIDES, "MINUTES": MINUTES,
        "USER_MASKS": USER_MASKS,
        "KM_PAYS_M": dict(KM_PAYS), "RIDES_PAYS_M": dict(RIDES_PAYS),
        "KM_MARQUE_M": dict(KM_MARQUE), "KM_VILLE_M": dict(KM_VILLE),
        "PAYS_CYCLISTES": dict(sorted(Counter(uid_pays.values()).items(), key=lambda kv: -kv[1])),
        "VILLE_CYCLISTES": dict(sorted(Counter(uid_ville.values()).items(), key=lambda kv: -kv[1])),
        "MARQUE_VELOS": dict(sorted(Counter(v["Marque"] for v in velos).items(), key=lambda kv: -kv[1])),
        # Ordre de 1re occurrence (pas de tri) : les maquettes cyclisme attendent
        # Depasse/Critique/OK/Alerte (ordre d'apparition dans FAIT_USURE_COMPOSANT).
        "USURE_STATUT": dict(Counter(u.get("Statut_Alerte") for u in usure if u.get("Statut_Alerte"))),
        "NB_UTILISATEURS": len(users), "NB_PAYS": len(set(uid_pays.values())),
        "NB_VILLES": len(set(uid_ville.values())), "NB_VELOS": len(velos),
        "NB_MARQUES": len(set(v["Marque"] for v in velos)),
        "ANNEE_MOY": round(sum(v["Annee_Sortie"] for v in velos) / len(velos), 1) if velos else 0,
        "USERS_AVEC_VELO": len({a["ID_Utilisateur"] for a in assoc}),
        "VELOS_ATTRIBUES": len(assigned),
    }
    return out


# ---------------------------------------------------------------------------
# Émission
# ---------------------------------------------------------------------------
def emit_normalized(data):
    return "const DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"


def emit_cyclisme(out):
    lines = []
    for k, v in out.items():
        lines.append("const %s = %s;" % (k, json.dumps(v, ensure_ascii=False)))
    return "\n".join(lines) + "\n"


def suggest_views(data):
    """Brouillon declaratif de views.json depuis le contrat DATA normalisé.

    Heuristique générique (tout domaine) :
      * une mesure -> KPI 'sum' (YoY) + visuel 'line' (N vs N-1)
      * une dimension -> KPI 'scalar' (NB_<dim>) + donut (<=6 valeurs) ou hbar (>6)
      * une source catégorielle (ex. usure) -> donut
      * l'entité active -> KPI 'active' (YoY)
    Le tout sur une page « Synthèse » unique, à raffiner par l'utilisateur.
    """
    meta = data["META"]
    measures = meta["measures"]
    dims = meta["dims"]
    kpis, visuals = [], []

    if meta.get("activity_entity"):
        kpis.append({"label": "Entités actives", "agg": "active", "yoy": True,
                     "fmt": "int", "sub": "actives en {CUR_YEAR}"})

    for m in measures:
        label = m.replace("_", " ")
        kpis.append({"label": "Total " + label, "agg": "sum", "measure": m,
                     "yoy": True, "fmt": "km" if "KM" in m.upper() else "int",
                     "sub": label + " en {CUR_YEAR}"})
        visuals.append({"type": "line", "measure": m,
                        "title": label.capitalize() + " par mois",
                        "unit": "km" if "KM" in m.upper() else "int"})

    for d in dims:
        dname = d["name"]
        nb_key = "NB_" + re.sub(r"\W+", "_", dname).upper()
        kpis.append({"label": dname + " (distincts)", "agg": "scalar",
                     "from": "SCALARS." + nb_key, "fmt": "int",
                     "sub": dname + " couverts"})
        n = len(data["DIM_COUNTS"].get(dname, {}))
        vtype = "donut" if n <= 6 else "hbar"
        visuals.append({"type": vtype, "from": "DIM_COUNTS." + dname, "top": 6 if vtype == "donut" else 10,
                        "title": "Par " + dname.lower(), "unit": "int",
                        "centerSub": dname.lower()})

    for src, cols in meta.get("category_sources", {}).items():
        for col in cols:
            visuals.append({"type": "donut",
                            "from": "CATEGORY_COUNTS.%s.%s" % (src, col), "top": 6,
                            "title": col.replace("_", " ").capitalize(),
                            "unit": "int", "centerSub": col.replace("_", " ").lower()})

    return {"labels": {}, "pages": [{"name": "Synthèse", "desc": "Vue d'ensemble auto-générée — à raffiner dans views.json.",
                                      "subs": [{"name": "Vue d'ensemble", "kpis": kpis, "visuals": visuals}]}]}


def main():
    ap = argparse.ArgumentParser(description="Extracteur donnees.xlsx -> bloc DATA JS")
    ap.add_argument("xlsx")
    ap.add_argument("--profile", choices=["cyclisme"], help="contrat legacy cyclisme (déprécié)")
    ap.add_argument("--manifest", help="chemin vers data-manifest.json (override)")
    ap.add_argument("--suggest-views", action="store_true",
                    help="émet un brouillon de views.json (JSON) au lieu du bloc DATA")
    args = ap.parse_args()

    sheets = load_wb(args.xlsx)
    if not sheets:
        sys.stderr.write("ERREUR: classeur vide ou illisible: %s\n" % args.xlsx)
        sys.exit(1)

    if args.profile == "cyclisme":
        out = build_cyclisme(sheets)
        sys.stdout.write(emit_cyclisme(out))
        sys.stderr.write("OK [cyclisme]: %d mois, %s mesures, %s usure\n" % (
            len(out["MONTH_META"]), len(out["KM"]), out["USURE_STATUT"]))
        return

    info = profile(sheets)
    manifest = None
    if args.manifest and os.path.exists(args.manifest):
        with open(args.manifest, encoding="utf-8") as f:
            manifest = json.load(f)
        sys.stderr.write("Manifeste chargé: %s\n" % args.manifest)

    try:
        data, proposed = build_normalized(sheets, info, manifest)
    except SystemExit as e:
        sys.stderr.write(str(e) + "\n")
        sys.exit(1)

    if args.suggest_views:
        sys.stdout.write(json.dumps(suggest_views(data), ensure_ascii=False, indent=2) + "\n")
        sys.stderr.write("OK: brouillon views.json émis (à raffiner).\n")
        return

    sys.stdout.write(emit_normalized(data))
    meta = data["META"]
    sys.stderr.write(
        "OK: faits='%s' date='%s' | %d mois | mesures=%s | dims=%s | entité_active=%s\n"
        % (meta["fact_sheet"], meta["date_col"], data["N"], meta["measures"],
           [d["name"] for d in meta["dims"]], meta["activity_entity"]))
    sys.stderr.write("--- Manifeste proposé (copiez dans data-manifest.json pour corriger) ---\n")
    sys.stderr.write(json.dumps(proposed, ensure_ascii=False, indent=2) + "\n")


if __name__ == "__main__":
    main()
